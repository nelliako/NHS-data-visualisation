import os, re
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Config 

CSV_DIR  = './csv_files'   # 1998-99 to 2011-12 summary CSVs
XLSX_DIR = './xlsx_files'  # 2012-13 to 2023-24 diagnosis xlsx files
OUT_PATH = './Chapter5_FCE.xlsx'

CHAPTERS = [
    ('F00-F09',  0,  9, 'Organic, including symptomatic, mental disorders'),
    ('F10-F19', 10, 19, 'Mental and behavioural disorders due to psychoactive substance use'),
    ('F20-F29', 20, 29, 'Schizophrenia, schizotypal and delusional disorders'),
    ('F30-F39', 30, 39, 'Mood [affective] disorders'),
    ('F40-F48', 40, 48, 'Neurotic, stress-related and somatoform disorders'),
    ('F50-F59', 50, 59, 'Behavioural syndromes associated with physiological disturbances'),
    ('F60-F69', 60, 69, 'Disorders of adult personality and behaviour'),
    ('F70-F79', 70, 79, 'Mental retardation'),
    ('F80-F89', 80, 89, 'Disorders of psychological development'),
    ('F90-F98', 90, 98, 'Behavioural and emotional disorders with onset in childhood'),
    ('F99-F99', 99, 99, 'Unspecified mental disorder'),
]


# Helpers 

def clean_num(val):
    s = re.sub(r'\(.*?\)', '', str(val)).replace(',', '').replace('*', '').replace('-', '').strip()
    try:
        return int(float(s)) if s else 0
    except:
        return 0

def get_year(fname):
    m = re.search(r'(\d{4})-(\d{2})-prim', fname)
    if m:
        return f"{m.group(1)}-{m.group(2)}"
    m = re.search(r'(\d{2})-(\d{2})', fname)
    if m:
        y1, y2 = m.group(1), m.group(2)
        return f"{'19' if int(y1) >= 98 else '20'}{y1}-{y2}"
    m = re.search(r'diag(\d{4})(\d{2})', fname)
    if m:
        return f"{m.group(1)}-{m.group(2)}"
    return None

# Extract from CSVs (1998-99 to 2011-12)

records = []

for fname in sorted(os.listdir(CSV_DIR)):
    if not fname.endswith('.csv'):
        continue
    year = get_year(fname)
    if not year:
        continue

    df = pd.read_csv(os.path.join(CSV_DIR, fname), header=None, on_bad_lines='skip')
    hdr = next(i for i, row in df.iterrows()
               if 'finished consultant' in ' '.join(str(x) for x in row.values).lower())
    fce_col = next(j for j, v in enumerate(df.iloc[hdr].astype(str).str.lower())
                   if 'finished consultant' in v)

    # Collect raw grouped rows
    raw = {}
    for i in range(hdr + 1, len(df)):
        desc = str(df.iloc[i, 0]).strip().upper()
        for prefix in ['F00-F03', 'F04-F09', 'F10-F19', 'F20-F29',
                       'F30-F39', 'F40-F69', 'F70-F79', 'F80-F99']:
            if desc.startswith(prefix):
                raw[prefix] = clean_num(df.iloc[i, fce_col])
                break

    f40_69 = raw.get('F40-F69')
    f80_99 = raw.get('F80-F99')

    vals = {
        'F00-F09': raw.get('F00-F03', 0) + raw.get('F04-F09', 0),
        'F10-F19': raw.get('F10-F19'),
        'F20-F29': raw.get('F20-F29'),
        'F30-F39': raw.get('F30-F39'),
        'F40-F48': f'incl. in F40-F69 ({f40_69:,})' if f40_69 else 'n/a',
        'F50-F59': 'incl. in F40-F69' if f40_69 else 'n/a',
        'F60-F69': 'incl. in F40-F69' if f40_69 else 'n/a',
        'F70-F79': raw.get('F70-F79'),
        'F80-F89': f'incl. in F80-F99 ({f80_99:,})' if f80_99 else 'n/a',
        'F90-F98': 'incl. in F80-F99' if f80_99 else 'n/a',
        'F99-F99': 'incl. in F80-F99' if f80_99 else 'n/a',
    }

    for key, _, _, desc in CHAPTERS:
        records.append({'Year': year, 'Chapter': key, 'Description': desc, 'Value': vals.get(key, 'n/a')})


# Extract from XLSXs (2012-13 to 2023-24) 

xlsx_files = sorted([f for f in os.listdir(XLSX_DIR)
                     if f.endswith('.xlsx') and 'diag' in f.lower()
                     and not any(x in f.lower() for x in ['4cha', 'sum', 'prim', 'dementia'])])

for fname in xlsx_files:
    year = get_year(fname)
    if not year:
        continue

    xl    = pd.ExcelFile(os.path.join(XLSX_DIR, fname))
    sheet = next(s for s in xl.sheet_names if 'All' in s and '3' in s)
    raw   = pd.read_excel(os.path.join(XLSX_DIR, fname), sheet_name=sheet, header=None, nrows=30)

    hdr      = next(i for i, row in raw.iterrows()
                    if 'Main diagnosis' in ' '.join(str(x) for x in row.values))
    hdr_vals = raw.iloc[hdr].astype(str).tolist()
    code_col = next(j for j, v in enumerate(hdr_vals) if 'all diagnoses: 3 character' in v.lower())
    main_col = next(j for j, v in enumerate(hdr_vals) if v.strip().lower() == 'main diagnosis')

    data = pd.read_excel(os.path.join(XLSX_DIR, fname), sheet_name=sheet, header=None, skiprows=hdr + 1)
    data['_code'] = data.iloc[:, code_col].astype(str).str.strip().str[:3].str.upper()
    data['_num']  = pd.to_numeric(data['_code'].str[1:], errors='coerce')
    data['_fce']  = data.iloc[:, main_col].apply(clean_num)
    fdata = data[data['_code'].str.match(r'^F\d{2}$')]

    for key, lo, hi, desc in CHAPTERS:
        total = int(fdata[(fdata['_num'] >= lo) & (fdata['_num'] <= hi)]['_fce'].sum())
        records.append({'Year': year, 'Chapter': key, 'Description': desc, 'Value': total})

# Pivot and write

df = pd.DataFrame(records)
years = sorted(df['Year'].unique(), key=lambda y: int(y[:4]))

pivot = []
for key, _, _, desc in CHAPTERS:
    row = {'Chapter': key, 'Description': desc}
    for yr in years:
        subset = df[(df['Year'] == yr) & (df['Chapter'] == key)]
        row[yr] = subset.iloc[0]['Value'] if not subset.empty else ''
    pivot.append(row)

pivot_df = pd.DataFrame(pivot)

wb = Workbook()
ws = wb.active
ws.title = 'Chapter 5 FCE'

headers = ['Chapter', 'Description'] + years
for ci, h in enumerate(headers, 1):
    ws.cell(row=1, column=ci, value=h)

for ri, row in pivot_df.iterrows():
    for ci, h in enumerate(headers, 1):
        val = row[h]
        cell = ws.cell(row=ri + 2, column=ci, value=val)
        if isinstance(val, int):
            cell.number_format = '#,##0'

ws.column_dimensions['A'].width = 12
ws.column_dimensions['B'].width = 52
for ci in range(3, 3 + len(years)):
    ws.column_dimensions[get_column_letter(ci)].width = 11

wb.save(OUT_PATH)
